"""
=============================================================================
GA_OBGYN_Rural_NPI_Lookup_v6.py
=============================================================================

README
------
PURPOSE
  Build a defensible public-data estimate of Georgia OB/GYN physicians
  from the full CMS/NPPES monthly data dissemination file.

WHERE TO PLACE THE NPPES ZIP
  Drop the NPPES_Data_Dissemination_*.zip file into the same folder as this script.

INSTALL DEPENDENCIES
  python -m pip install pandas duckdb pyarrow openpyxl requests tqdm

RUN THE SCRIPT
  python GA_OBGYN_Rural_NPI_Lookup_v6.py

REFRESH PROCESSED DATA
  Delete or rename the files in the processed/ subfolder,
  then re-run. Or answer "y" when prompted.

OUTPUTS
  outputs/GA_OBGYN_Rural_v6_CLEAN.xlsx
  outputs/GA_OBGYN_Rural_v6_AUDIT.xlsx
  outputs/GA_OBGYN_Rural_v6_Core_General.csv
  outputs/GA_OBGYN_Rural_v6_Core_Umbrella.csv
  outputs/GA_OBGYN_Rural_v6_All_Provider_Audit.csv

NOTE
  This is an NPI/NPPES-derived evidence-weighted estimate, not a census.
  See Methodology Notes tab in the clean workbook for full limitations.
=============================================================================
"""

# ---------------------------------------------------------------------------
# 0.  Dependency check
# ---------------------------------------------------------------------------
import sys
import importlib

REQUIRED = {
    "pandas": "pandas",
    "duckdb": "duckdb",
    "pyarrow": "pyarrow",
    "openpyxl": "openpyxl",
    "requests": "requests",
    "tqdm": "tqdm",
}
missing = []
for pkg, import_name in REQUIRED.items():
    try:
        importlib.import_module(import_name)
    except ImportError:
        missing.append(pkg)

if missing:
    print("ERROR: Missing required packages.")
    print("Run:  python -m pip install " + " ".join(missing))
    sys.exit(1)

import zipfile
import json
import time
import re
import os
import collections
from pathlib import Path
from datetime import date

import pandas as pd
import duckdb
import requests
from tqdm import tqdm

# ---------------------------------------------------------------------------
# 1.  Configuration — edit paths and dictionaries here
# ---------------------------------------------------------------------------

BASE_DIR   = Path(__file__).parent
RAW_DIR    = BASE_DIR / "raw"
PROC_DIR   = BASE_DIR / "processed"
OUT_DIR    = BASE_DIR / "outputs"

GEOCODE_CACHE_PATH = PROC_DIR / "geocode_cache_ga_obgyn_v6.json"
DUCKDB_PATH        = PROC_DIR / "ga_obgyn_v6.duckdb"
PARQUET_PROVIDERS  = PROC_DIR / "ga_obgyn_candidates.parquet"
PARQUET_LOCATIONS  = PROC_DIR / "ga_obgyn_locations.parquet"

CHUNK_SIZE = 100_000   # rows per chunk when reading main NPPES CSV

# 207V OB/GYN taxonomy codes
OBGYN_TAXONOMY_CODES = {
    "207V00000X": "Obstetrics & Gynecology",
    "207VB0002X": "Obesity Medicine (OB/GYN)",
    "207VC0200X": "Critical Care Medicine (OB/GYN)",
    "207VC0300X": "Complex Family Planning",
    "207VE0102X": "Reproductive Endocrinology",
    "207VF0040X": "Female Pelvic Medicine & Reconstructive Surgery",
    "207VG0400X": "Gynecology",
    "207VH0002X": "Hospice & Palliative Medicine (OB/GYN)",
    "207VM0101X": "Maternal & Fetal Medicine",
    "207VX0000X": "Obstetrics",
    "207VX0201X": "Gynecologic Oncology",
}

GENERAL_OBGYN_CODE = "207V00000X"

# Taxonomy codes that are clearly subspecialist-only (not general OB/GYN)
SUBSPECIALIST_ONLY_CODES = {
    "207VB0002X", "207VC0200X", "207VC0300X", "207VE0102X",
    "207VF0040X", "207VH0002X", "207VM0101X", "207VX0201X",
}

# ---------------------------------------------------------------------------
# Georgia county rural classification
# Edit this dictionary to reflect your rural definition.
# Values: "Rural" | "Not rural" | "Partial rural - tract/address review needed"
# ---------------------------------------------------------------------------
GA_COUNTY_RURAL = {
    "Appling": "Rural",
    "Atkinson": "Rural",
    "Bacon": "Rural",
    "Baker": "Rural",
    "Baldwin": "Partial rural - tract/address review needed",
    "Banks": "Rural",
    "Barrow": "Not rural",
    "Bartow": "Partial rural - tract/address review needed",
    "Ben Hill": "Rural",
    "Berrien": "Rural",
    "Bibb": "Not rural",
    "Bleckley": "Rural",
    "Brantley": "Rural",
    "Brooks": "Rural",
    "Bryan": "Partial rural - tract/address review needed",
    "Bulloch": "Partial rural - tract/address review needed",
    "Burke": "Rural",
    "Butts": "Rural",
    "Calhoun": "Rural",
    "Camden": "Partial rural - tract/address review needed",
    "Candler": "Rural",
    "Carroll": "Partial rural - tract/address review needed",
    "Catoosa": "Partial rural - tract/address review needed",
    "Charlton": "Rural",
    "Chatham": "Not rural",
    "Chattahoochee": "Rural",
    "Chattooga": "Rural",
    "Cherokee": "Not rural",
    "Clarke": "Not rural",
    "Clay": "Rural",
    "Clayton": "Not rural",
    "Clinch": "Rural",
    "Cobb": "Not rural",
    "Coffee": "Rural",
    "Colquitt": "Rural",
    "Columbia": "Not rural",
    "Cook": "Rural",
    "Coweta": "Not rural",
    "Crawford": "Rural",
    "Crisp": "Rural",
    "Dade": "Rural",
    "Dawson": "Partial rural - tract/address review needed",
    "Decatur": "Rural",
    "DeKalb": "Not rural",
    "Dodge": "Rural",
    "Dooly": "Rural",
    "Dougherty": "Partial rural - tract/address review needed",
    "Douglas": "Not rural",
    "Early": "Rural",
    "Echols": "Rural",
    "Effingham": "Partial rural - tract/address review needed",
    "Elbert": "Rural",
    "Emanuel": "Rural",
    "Evans": "Rural",
    "Fannin": "Rural",
    "Fayette": "Not rural",
    "Floyd": "Partial rural - tract/address review needed",
    "Forsyth": "Not rural",
    "Franklin": "Rural",
    "Fulton": "Not rural",
    "Gilmer": "Rural",
    "Glascock": "Rural",
    "Glynn": "Partial rural - tract/address review needed",
    "Gordon": "Partial rural - tract/address review needed",
    "Grady": "Rural",
    "Greene": "Rural",
    "Gwinnett": "Not rural",
    "Habersham": "Rural",
    "Hall": "Not rural",
    "Hancock": "Rural",
    "Haralson": "Rural",
    "Harris": "Partial rural - tract/address review needed",
    "Hart": "Rural",
    "Heard": "Rural",
    "Henry": "Not rural",
    "Houston": "Not rural",
    "Irwin": "Rural",
    "Jackson": "Partial rural - tract/address review needed",
    "Jasper": "Rural",
    "Jeff Davis": "Rural",
    "Jefferson": "Rural",
    "Jenkins": "Rural",
    "Johnson": "Rural",
    "Jones": "Rural",
    "Lamar": "Rural",
    "Lanier": "Rural",
    "Laurens": "Rural",
    "Lee": "Partial rural - tract/address review needed",
    "Liberty": "Partial rural - tract/address review needed",
    "Lincoln": "Rural",
    "Long": "Rural",
    "Lowndes": "Partial rural - tract/address review needed",
    "Lumpkin": "Rural",
    "McDuffie": "Rural",
    "McIntosh": "Rural",
    "Macon": "Rural",
    "Madison": "Rural",
    "Marion": "Rural",
    "Meriwether": "Rural",
    "Miller": "Rural",
    "Mitchell": "Rural",
    "Monroe": "Rural",
    "Montgomery": "Rural",
    "Morgan": "Rural",
    "Murray": "Rural",
    "Muscogee": "Not rural",
    "Newton": "Partial rural - tract/address review needed",
    "Oconee": "Partial rural - tract/address review needed",
    "Oglethorpe": "Rural",
    "Paulding": "Not rural",
    "Peach": "Rural",
    "Pickens": "Partial rural - tract/address review needed",
    "Pierce": "Rural",
    "Pike": "Rural",
    "Polk": "Rural",
    "Pulaski": "Rural",
    "Putnam": "Rural",
    "Quitman": "Rural",
    "Rabun": "Rural",
    "Randolph": "Rural",
    "Richmond": "Not rural",
    "Rockdale": "Not rural",
    "Schley": "Rural",
    "Screven": "Rural",
    "Seminole": "Rural",
    "Spalding": "Partial rural - tract/address review needed",
    "Stephens": "Rural",
    "Stewart": "Rural",
    "Sumter": "Rural",
    "Talbot": "Rural",
    "Taliaferro": "Rural",
    "Tattnall": "Rural",
    "Taylor": "Rural",
    "Telfair": "Rural",
    "Terrell": "Rural",
    "Thomas": "Partial rural - tract/address review needed",
    "Tift": "Rural",
    "Toombs": "Rural",
    "Towns": "Rural",
    "Treutlen": "Rural",
    "Troup": "Partial rural - tract/address review needed",
    "Turner": "Rural",
    "Twiggs": "Rural",
    "Union": "Rural",
    "Upson": "Rural",
    "Walker": "Partial rural - tract/address review needed",
    "Walton": "Partial rural - tract/address review needed",
    "Ware": "Rural",
    "Warren": "Rural",
    "Washington": "Rural",
    "Wayne": "Rural",
    "Webster": "Rural",
    "Wheeler": "Rural",
    "White": "Rural",
    "Whitfield": "Partial rural - tract/address review needed",
    "Wilcox": "Rural",
    "Wilkes": "Rural",
    "Wilkinson": "Rural",
    "Worth": "Rural",
}

# ---------------------------------------------------------------------------
# Workforce exclusion rules — aligned with HRSA HWSM methodology
# HRSA definition: "active in the workforce providing direct patient care;
# excludes residents and physicians in non-patient-care settings."
# ---------------------------------------------------------------------------

# Non-physician credential patterns — these are advanced practice providers,
# not physicians. Valuable to the OB/GYN workforce but excluded from the
# physician-only count to match HRSA methodology.
NON_PHYSICIAN_CREDENTIAL_PATTERNS = re.compile(
    r"\b(CNM|CNP|WHNP|FNP|NP|PA-C|PAC|PA\b|APRN|RN\b|MSN|DNP|RDMS|CRNA)\b",
    re.IGNORECASE,
)

# Non-OB/GYN taxonomy codes that should not appear in OB/GYN counts.
# These slipped through because they co-occur with 207V codes on the same NPI.
NON_OBGYN_PRIMARY_TAXONOMIES = {
    "207Q00000X",   # Family Medicine
    "2086S0102X",   # Surgical Oncology
    "174400000X",   # Health & Wellness Coach
    "176B00000X",   # Midwife (not physician)
    "367A00000X",   # Advanced Practice Midwife
    "363LA2200X",   # Adult Health NP
    "363LW0102X",   # Women's Health NP
    "363LF0000X",   # Family NP
    "207RE0101X",   # Reproductive Endocrinology (PA)
    "207VX0000X",   # Obstetrics — keep if physician, exclude if CNM/NP
}

# Known Georgia OB/GYN residency and fellowship training program addresses.
# Providers at these addresses without credentials or with mailstop-style
# room numbers are flagged as likely residents/trainees, not attending physicians.
# Source: program websites + ACGME directory.
RESIDENCY_ADDRESSES = {
    # Augusta University / MCG (AU Health)
    "1120 15TH ST",
    "1120 15TH STREET",
    "1459 LANEY WALKER BLVD",
    # Emory University (Grady, Emory Midtown, Emory University Hospital)
    "80 JESSE HILL JR DR SE",
    "80 JESSE HILL JR DRIVE",
    "69 JESSE HILL JR DR SE",
    "1364 CLIFTON RD NE",
    "1365 CLIFTON RD NE",
    # Mercer University / Medical Center Navicent Health (Macon)
    "777 HEMLOCK ST",
    "777 HEMLOCK STREET",
    # Morehouse School of Medicine / Grady
    "720 WESTVIEW DR SW",
    # Medical College of Georgia / Fort Gordon
    "DWIGHT D. EISENHOWER ARMY MEDICAL CENTER",
    # Uniformed Services / Fort Benning
    "6600 VAN AALST BLVD",
    # Department-only address (no real street)
    "DEPARTMENT OF OBSTETRICS & GYNECOLOGY",
    "DEPARTMENT OF OBSTETRICS AND GYNECOLOGY",
}

# Mailstop / room-number patterns used by residency programs
# (e.g., "# 165", "MSC:167", "BB-1234") — attending physicians use suite numbers
RESIDENT_ROOM_PATTERN = re.compile(
    r"(MSC\s*:\s*\d+|#\s*\d{3,}(?!\s*(ST|AVE|RD|BLVD|DR|LN|WAY|PKW|HWY)))",
    re.IGNORECASE,
)


def is_likely_resident(address: str, credential: str, address_line2: str = "") -> bool:
    """
    Flag providers likely to be residents/trainees rather than attending physicians.
    Uses three signals aligned with HRSA's exclusion criteria:
      1. Address matches a known residency training site
      2. No physician credential (MD/DO) listed
      3. Address contains a mailstop or room number pattern
    Requires at least two of three signals to flag — avoids over-exclusion.
    """
    addr_upper = str(address).strip().upper()
    cred_upper = str(credential).strip().upper()

    # Check address against known residency sites
    at_residency_site = any(
        site in addr_upper for site in RESIDENCY_ADDRESSES
    )

    # No physician credential — blank, or non-MD/DO
    has_physician_cred = bool(re.search(r"\b(M\.?D\.?|D\.?O\.?)\b", cred_upper))
    no_physician_cred  = not has_physician_cred and cred_upper in ("", "NAN", "NONE")

    # Mailstop / room number pattern (residents use mailstops, attendings use suites)
    has_mailstop = bool(RESIDENT_ROOM_PATTERN.search(addr_upper))

    # Score: need 2+ signals
    score = int(at_residency_site) + int(no_physician_cred) + int(has_mailstop)
    return score >= 2


def is_non_physician(credential: str, primary_taxonomy: str) -> bool:
    """
    Returns True if the provider is a non-physician (CNM, NP, PA, RN, etc.).
    These are valuable OB/GYN workforce members but excluded from the
    physician-only count per HRSA methodology.
    """
    cred = str(credential).strip()
    if NON_PHYSICIAN_CREDENTIAL_PATTERNS.search(cred):
        return True
    # Also exclude by primary taxonomy if it is a non-physician code
    tax = str(primary_taxonomy).strip()
    if tax in NON_OBGYN_PRIMARY_TAXONOMIES - {"207VX0000X"}:
        # 207VX0000X (Obstetrics) is valid for physicians — only exclude
        # if credential is also non-physician
        return True
    return False


# ---------------------------------------------------------------------------
# 2.  Utility helpers
# ---------------------------------------------------------------------------

def make_dirs():
    for d in [RAW_DIR, PROC_DIR, OUT_DIR]:
        d.mkdir(parents=True, exist_ok=True)

def find_nppes_zip() -> Path:
    """Find the NPPES ZIP in BASE_DIR."""
    zips = sorted(BASE_DIR.glob("NPPES_Data_Dissemination*.zip"))
    if not zips:
        zips = sorted(BASE_DIR.glob("*.zip"))
    if not zips:
        raise FileNotFoundError(
            f"No NPPES ZIP found in {BASE_DIR}.\n"
            "Download from https://download.cms.gov/nppes/NPI_Files.html "
            "and place the ZIP in that folder."
        )
    if len(zips) > 1:
        print(f"  [WARN] Multiple ZIPs found; using newest: {zips[-1].name}")
    return zips[-1]


def list_zip_contents(zip_path: Path) -> list:
    with zipfile.ZipFile(zip_path, "r") as zf:
        return zf.namelist()


def find_main_nppes_csv(names: list) -> str:
    """Identify the main NPPES provider file inside the ZIP."""
    # Main file is usually named npidata_pfile_YYYYMMDD-YYYYMMDD.csv
    for n in names:
        base = n.lower()
        if "npidata_pfile" in base and base.endswith(".csv") and "fileheader" not in base:
            return n
    # Fallback: largest CSV
    csv_names = [n for n in names if n.lower().endswith(".csv") and "fileheader" not in n.lower()]
    return csv_names[0] if csv_names else None


def find_pl_csv(names: list) -> str:
    """Identify the Practice Location Reference File."""
    for n in names:
        base = n.lower()
        if ("pl_pfile" in base or "othername" in base or "endpoint" in base
                or "practicelocation" in base) and base.endswith(".csv"):
            return n
    return None


def normalize_address(street: str, city: str, state: str, zipcode: str) -> str:
    """Create a normalized cache key from address components."""
    parts = [str(x).strip().upper() for x in [street, city, state, zipcode] if str(x).strip()]
    return "|".join(parts)


def load_geocode_cache() -> dict:
    if GEOCODE_CACHE_PATH.exists():
        try:
            with open(GEOCODE_CACHE_PATH, "r") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_geocode_cache(cache: dict):
    with open(GEOCODE_CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)


# ---------------------------------------------------------------------------
# Georgia ZIP-to-county fallback table
# Covers military bases, hospitals, and common geocoder failures.
# Format: "ZIP": "County Name" (must match GA_COUNTY_RURAL keys exactly)
# ---------------------------------------------------------------------------
GA_ZIP_TO_COUNTY = {
    # Military bases
    "31314": "Liberty",       # Fort Stewart
    "31315": "Liberty",       # Fort Stewart
    "31905": "Muscogee",      # Fort Benning (now Fort Moore)
    "31908": "Muscogee",      # Fort Benning
    "30905": "Richmond",      # Fort Gordon (now Fort Eisenhower)
    "30906": "Richmond",      # Augusta / Fort Gordon area
    # Augusta / Richmond County
    "30901": "Richmond",
    "30904": "Richmond",
    "30907": "Columbia",
    "30909": "Richmond",
    # Atlanta metro
    "30309": "Fulton",
    "30310": "Fulton",
    "30311": "Fulton",
    "30312": "Fulton",
    "30313": "Fulton",
    "30314": "Fulton",
    "30315": "Fulton",
    "30316": "DeKalb",
    "30317": "DeKalb",
    "30318": "Fulton",
    "30319": "DeKalb",
    "30324": "DeKalb",
    "30326": "Fulton",
    "30327": "Fulton",
    "30328": "Fulton",
    "30329": "DeKalb",
    "30331": "Fulton",
    "30332": "Fulton",        # Georgia Tech
    "30333": "DeKalb",
    "30338": "DeKalb",
    "30339": "Cobb",
    "30340": "DeKalb",
    "30341": "DeKalb",
    "30342": "Fulton",
    "30345": "DeKalb",
    "30346": "DeKalb",
    "30360": "DeKalb",
    # Emory / Decatur area
    "30030": "DeKalb",
    "30032": "DeKalb",
    "30033": "DeKalb",
    "30034": "DeKalb",
    "30035": "DeKalb",
    "30058": "DeKalb",        # Lithonia
    "30079": "DeKalb",
    "30083": "DeKalb",
    "30084": "DeKalb",        # Tucker
    "30322": "DeKalb",        # Emory University Hospital
    # Cobb County
    "30060": "Cobb",
    "30062": "Cobb",
    "30064": "Cobb",
    "30066": "Cobb",
    "30067": "Cobb",
    "30068": "Cobb",
    "30080": "Cobb",
    "30101": "Cherokee",
    # Gwinnett
    "30024": "Gwinnett",
    "30040": "Forsyth",
    "30041": "Forsyth",
    "30043": "Gwinnett",
    "30044": "Gwinnett",
    "30045": "Gwinnett",
    "30046": "Gwinnett",
    "30047": "Gwinnett",
    "30096": "Gwinnett",
    "30097": "Gwinnett",
    # Savannah / Chatham
    "31401": "Chatham",
    "31404": "Chatham",
    "31405": "Chatham",
    "31406": "Chatham",
    "31407": "Chatham",
    "31408": "Chatham",
    "31419": "Chatham",
    # Columbus / Muscogee
    "31901": "Muscogee",
    "31902": "Muscogee",
    "31903": "Muscogee",
    "31904": "Muscogee",
    "31906": "Muscogee",
    "31907": "Muscogee",
    "31909": "Muscogee",
    # Macon / Bibb
    "31201": "Bibb",
    "31204": "Bibb",
    "31206": "Bibb",
    "31210": "Bibb",
    "31211": "Bibb",
    "31216": "Bibb",
    "31217": "Bibb",
    "31220": "Bibb",
    # Warner Robins / Houston
    "31088": "Houston",
    "31093": "Houston",
    "31098": "Houston",
    # Albany / Dougherty
    "31701": "Dougherty",
    "31705": "Dougherty",
    "31707": "Dougherty",
    # Gainesville / Hall
    "30501": "Hall",
    "30504": "Hall",
    "30506": "Hall",
    "30507": "Hall",
    # Athens / Clarke
    "30601": "Clarke",
    "30605": "Clarke",
    "30606": "Clarke",
    "30607": "Clarke",
    # Valdosta / Lowndes
    "31601": "Lowndes",
    "31602": "Lowndes",
    "31605": "Lowndes",
    # Brunswick / Glynn
    "31520": "Glynn",
    "31523": "Glynn",
    "31525": "Glynn",
    # Rome / Floyd
    "30161": "Floyd",
    "30165": "Floyd",
    # Dalton / Whitfield
    "30720": "Whitfield",
    "30721": "Whitfield",
    # Douglasville / Douglas
    "30134": "Douglas",
    "30135": "Douglas",
    # Woodstock / Cherokee
    "30188": "Cherokee",
    "30189": "Cherokee",
    # Carrollton / Carroll
    "30116": "Carroll",
    "30117": "Carroll",
    # Newnan / Coweta
    "30263": "Coweta",
    "30265": "Coweta",
    # LaGrange / Troup
    "30240": "Troup",
    "30241": "Troup",
    # Statesboro / Bulloch
    "30458": "Bulloch",
    "30461": "Bulloch",
    # Waycross / Ware
    "31501": "Ware",
    "31503": "Ware",
    # Thomasville / Thomas
    "31792": "Thomas",
    "31799": "Thomas",
    # Hinesville / Liberty
    "31313": "Liberty",
    # McDonough / Henry
    "30252": "Henry",
    "30253": "Henry",
    # Stockbridge / Henry
    "30281": "Henry",
    # Smyrna / Cobb
    "30080": "Cobb",
    "30082": "Cobb",
    # Villa Rica / Carroll
    "30180": "Carroll",
    # Blairsville / Union
    "30512": "Union",
    # Clayton / Rabun
    "30525": "Rabun",
    # Blue Ridge / Fannin
    "30513": "Fannin",
    # Cumming / Forsyth
    "30028": "Forsyth",
    # Jasper / Pickens
    "30143": "Pickens",
    # Canton / Cherokee
    "30114": "Cherokee",
    "30115": "Cherokee",
    # Alpharetta / Fulton
    "30004": "Fulton",
    "30005": "Fulton",
    "30009": "Fulton",
    # Roswell / Fulton
    "30075": "Fulton",
    "30076": "Fulton",
    # Sandy Springs / Fulton
    "30350": "Fulton",
    "30358": "Fulton",
    # Marietta / Cobb
    "30008": "Cobb",
    "30060": "Cobb",
    "30061": "Cobb",
    "30062": "Cobb",
    "30063": "Cobb",
    "30064": "Cobb",
    "30065": "Cobb",
    "30066": "Cobb",
    "30067": "Cobb",
    "30068": "Cobb",
    "30069": "Cobb",
    # Kennesaw / Cobb
    "30144": "Cobb",
    "30152": "Cobb",
    # Acworth / Cobb
    "30101": "Cobb",
    "30102": "Cobb",
    # Fayetteville / Fayette
    "30214": "Fayette",
    "30215": "Fayette",
    "30269": "Fayette",      # Peachtree City
    # Moultrie / Colquitt
    "31768": "Colquitt",
    "31769": "Colquitt",
    # Tifton / Tift
    "31793": "Tift",
    "31794": "Tift",
    # Dublin / Laurens
    "31021": "Laurens",
    "31040": "Laurens",
    # Eastman / Dodge
    "31023": "Dodge",
    # Brunswick / Glynn (additional ZIPs)
    "31522": "Glynn",
    "31524": "Glynn",
    # Lavonia / Franklin
    "30553": "Franklin",
    # Milledgeville / Baldwin
    "31061": "Baldwin",
    "31062": "Baldwin",
    # Zebulon / Pike
    "30295": "Pike",
    # Dawsonville / Dawson
    "30534": "Dawson",
    # Cartersville / Bartow
    "30120": "Bartow",
    "30121": "Bartow",
    # Jesup / Wayne
    "31545": "Wayne",
    "31546": "Wayne",
    # Fort Valley / Peach
    "31030": "Peach",
    # Sharpsburg / Coweta
    "30277": "Coweta",
    # Demorest / Habersham
    "30535": "Habersham",
    # Sautee Nacoochee / White
    "30571": "White",
    # Millen / Jenkins
    "30442": "Jenkins",
    # Vidalia / Toombs
    "30474": "Toombs",
    "30475": "Toombs",
    # Swainsboro / Emanuel
    "30401": "Emanuel",
    # Baxley / Appling
    "31513": "Appling",
    # Fitzgerald / Ben Hill
    "31750": "Ben Hill",
    # Douglas / Coffee
    "31533": "Coffee",
    "31535": "Coffee",
    # Cordele / Crisp
    "31010": "Crisp",
    # Americus / Sumter
    "31709": "Sumter",
    "31719": "Sumter",
    # Cairo / Grady
    "39827": "Grady",
    # Camilla / Mitchell
    "31730": "Mitchell",
    # Thomaston / Upson
    "30286": "Upson",
    # Forsyth / Monroe
    "31029": "Monroe",
    # Jackson / Butts
    "30233": "Butts",
    # Griffin / Spalding
    "30223": "Spalding",
    "30224": "Spalding",
    # Milledgeville / Baldwin (alias)
    "31059": "Baldwin",
    # Eatonton / Putnam
    "31024": "Putnam",
    # Greensboro / Greene
    "30642": "Greene",
    # Madison / Morgan
    "30650": "Morgan",
    # Covington / Newton
    "30014": "Newton",
    "30016": "Newton",
    # Monroe / Walton
    "30655": "Walton",
    "30656": "Walton",
    # Hartwell / Hart
    "30643": "Hart",
    # Elberton / Elbert
    "30635": "Elbert",
    # Washington / Wilkes
    "30673": "Wilkes",
    # Thomson / McDuffie
    "30824": "McDuffie",
    # Sandersville / Washington
    "31082": "Washington",
    # Louisville / Jefferson
    "30434": "Jefferson",
    # Sylvania / Screven
    "30467": "Screven",
    # Claxton / Evans
    "30417": "Evans",
    # Reidsville / Tattnall
    "30453": "Tattnall",
    # Hazlehurst / Jeff Davis
    "31539": "Jeff Davis",
    # Alma / Bacon
    "31510": "Bacon",
    # Homerville / Clinch
    "31634": "Clinch",
    # Folkston / Charlton
    "31537": "Charlton",
    # Nahunta / Brantley
    "31553": "Brantley",
    # Blackshear / Pierce
    "31516": "Pierce",
    # Patterson / Pierce
    "31557": "Pierce",
    # Quitman / Brooks
    "31643": "Brooks",
    # Lowndes (Valdosta area additional)
    "31606": "Lowndes",
    "31699": "Lowndes",
    # Adel / Cook
    "31620": "Cook",
    # Ashburn / Turner
    "31714": "Turner",
    # Sylvester / Worth
    "31791": "Worth",
    # Pelham / Mitchell
    "31779": "Mitchell",
    # Bainbridge / Decatur County (not DeKalb)
    "39817": "Decatur",
    "39819": "Decatur",
    # Blakely / Early
    "39823": "Early",
    # Cuthbert / Randolph
    "39840": "Randolph",
    # Georgetown / Quitman County
    "39854": "Quitman",
    # Lumpkin / Stewart
    "31815": "Stewart",
    # Buena Vista / Marion
    "31803": "Marion",
    # Butler / Taylor
    "31006": "Taylor",
    # Reynolds / Taylor
    "31076": "Taylor",
    # Vienna / Dooly
    "31092": "Dooly",
    # Hawkinsville / Pulaski
    "31036": "Pulaski",
    # McRae / Telfair
    "31055": "Telfair",
    # Abbeville / Wilcox
    "31001": "Wilcox",
    # Rochelle / Wilcox
    "31079": "Wilcox",
    # Ocilla / Irwin
    "31774": "Irwin",
    # Lakeland / Lanier
    "31635": "Lanier",
    # Nahunta (Brantley alias)
    "31563": "Brantley",
    # Chatsworth / Murray
    "30705": "Murray",
    # Ellijay / Gilmer
    "30536": "Gilmer",
    "30540": "Gilmer",
    # Hiawassee / Towns
    "30546": "Towns",
    # Young Harris / Towns
    "30582": "Towns",
    # Dahlonega / Lumpkin
    "30533": "Lumpkin",
    # Cleveland / White
    "30528": "White",
    # Helen / White
    "30545": "White",
    # Cornelia / Habersham
    "30531": "Habersham",
    # Toccoa / Stephens
    "30577": "Stephens",
    # Elberton / Elbert (alias)
    "30636": "Elbert",
    # Lincolnton / Lincoln
    "30817": "Lincoln",
    # Waynesboro / Burke
    "30830": "Burke",
    # Wrens / Jefferson
    "30833": "Jefferson",
    # Sparta / Hancock
    "31087": "Hancock",
    # Gray / Jones
    "31032": "Jones",
    # Forsyth / Monroe (alias)
    "31030": "Peach",        # Fort Valley / Peach (already above)
    # Barnesville / Lamar
    "30204": "Lamar",
    # Thomasville / Thomas (additional)
    "31757": "Thomas",
    # Moultrie / Colquitt (alias)
    "31776": "Colquitt",
    # Donalsonville / Seminole
    "39845": "Seminole",
    # Colquitt / Miller County
    "39837": "Miller",
    # Calhoun / Gordon
    "30701": "Gordon",
    "30703": "Gordon",
    # La Fayette / Walker
    "30728": "Walker",
    # Summerville / Chattooga
    "30747": "Chattooga",
    # Cedartown / Polk
    "30125": "Polk",
    "30127": "Polk",
    # Bremen / Haralson
    "30110": "Haralson",
    # Tallapoosa / Haralson
    "30176": "Haralson",
    # Winder / Barrow
    "30680": "Barrow",
    "30681": "Barrow",
    # Jefferson / Jackson County
    "30549": "Jackson",
    # Commerce / Jackson County
    "30529": "Jackson",
    # Toccoa Falls / Stephens
    "30598": "Stephens",
    # Clarkesville / Habersham
    "30523": "Habersham",
    # Gainesville / Hall (additional)
    "30503": "Hall",
    "30504": "Hall",
    "30506": "Hall",
    "30507": "Hall",
    "30566": "Hall",
    # Oakwood / Hall
    "30566": "Hall",
}


def geocode_census(street: str, city: str, state: str, zipcode: str) -> dict:
    """
    Call U.S. Census Geocoder (batch-friendly single-address endpoint).
    Returns dict with keys: matched, county, lat, lon, match_type, matched_address
    """
    url = "https://geocoding.geo.census.gov/geocoder/geographies/address"
    params = {
        "street": street,
        "city": city,
        "state": state,
        "zip": zipcode,
        "benchmark": "Public_AR_Current",
        "vintage": "Current_Current",
        "layers": "Counties",
        "format": "json",
    }
    try:
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        data = r.json()
        matches = data.get("result", {}).get("addressMatches", [])
        if not matches:
            return {"matched": False, "county": None, "lat": None, "lon": None,
                    "match_type": "No match", "matched_address": None}
        m = matches[0]
        coords = m.get("coordinates", {})
        county_name = None
        for geo_key in ["Counties", "counties"]:
            geo_list = m.get("geographies", {}).get(geo_key, [])
            if geo_list:
                county_name = geo_list[0].get("NAME", None)
                break
        return {
            "matched": True,
            "county": county_name,
            "lat": coords.get("y"),
            "lon": coords.get("x"),
            "match_type": m.get("matchedAddress", ""),
            "matched_address": m.get("matchedAddress", ""),
        }
    except Exception as e:
        return {"matched": False, "county": None, "lat": None, "lon": None,
                "match_type": f"Error: {e}", "matched_address": None}


# ZIP-to-county fallback (unambiguous only — expand as needed)
ZIP_TO_COUNTY = {}  # populated lazily from geocode results


def is_po_box(address: str) -> bool:
    if not address:
        return False
    return bool(re.search(r"\bP\.?\s*O\.?\s*BOX\b", str(address).upper()))


def classify_rural(county: str) -> str:
    if not county or str(county).strip() == "":
        return "Unknown - geocode/manual review"
    # Normalize: strip "County" suffix if present
    c = re.sub(r"\s+County$", "", str(county).strip(), flags=re.IGNORECASE).strip()
    # Case-insensitive lookup — handles "DeKalb" vs "Dekalb" etc.
    c_lower = c.lower()
    for key, val in GA_COUNTY_RURAL.items():
        if key.lower() == c_lower:
            return val
    return "Unknown - geocode/manual review"


# ---------------------------------------------------------------------------
# 3.  Taxonomy column helpers for the wide NPPES layout
# ---------------------------------------------------------------------------

def build_taxonomy_cols(columns: list) -> list:
    """Return list of (taxonomy_code_col, taxonomy_switch_col) pairs."""
    pairs = []
    for col in columns:
        m = re.match(r"Healthcare Provider Taxonomy Code_(\d+)", col, re.IGNORECASE)
        if m:
            idx = m.group(1)
            sw_col = f"Healthcare Provider Primary Taxonomy Switch_{idx}"
            pairs.append((col, sw_col))
    return pairs


def extract_taxonomies(row: dict, tax_pairs: list) -> list:
    """Return list of (code, is_primary) for all non-empty taxonomy fields."""
    result = []
    for code_col, sw_col in tax_pairs:
        code = str(row.get(code_col, "") or "").strip()
        if code:
            is_primary = str(row.get(sw_col, "") or "").strip().upper() == "Y"
            result.append((code, is_primary))
    return result


def build_license_cols(columns: list) -> list:
    """Return list of (license_number_col, license_state_col) pairs."""
    pairs = []
    for col in columns:
        m = re.match(r"Provider License Number_(\d+)", col, re.IGNORECASE)
        if m:
            idx = m.group(1)
            state_col = f"Provider License Number State Code_{idx}"
            pairs.append((col, state_col))
    return pairs


def has_ga_license(row: dict, lic_pairs: list) -> bool:
    for num_col, state_col in lic_pairs:
        state = str(row.get(state_col, "") or "").strip().upper()
        if state == "GA":
            return True
    return False


def ga_license_on_obgyn(row: dict, lic_pairs: list, taxonomies: list) -> bool:
    """True if any GA license number column accompanies an OB/GYN taxonomy."""
    # NPPES doesn't directly link licenses to taxonomies,
    # so we proxy: provider has GA license AND OB/GYN taxonomy present.
    obgyn_present = any(code in OBGYN_TAXONOMY_CODES for code, _ in taxonomies)
    return obgyn_present and has_ga_license(row, lic_pairs)


# ---------------------------------------------------------------------------
# 4.  Confidence scoring
# ---------------------------------------------------------------------------

def compute_confidence(
    has_ga_primary_practice: bool,
    has_ga_secondary_practice: bool,
    ga_license_on_obgyn_tax: bool,
    has_any_ga_license: bool,
    has_primary_obgyn_tax: bool,
    has_general_obgyn_code: bool,
    clean_county_geocode: bool,
    not_po_box: bool,
    mailing_only: bool,
    failed_geocode: bool,
    is_po_box_addr: bool,
    multiple_conflicting_states: bool,
    oos_practice_with_ga_license: bool,
) -> int:
    score = 0
    if has_ga_primary_practice:
        score += 35
    if has_ga_secondary_practice:
        score += 25
    if ga_license_on_obgyn_tax:
        score += 25
    elif has_any_ga_license:
        score += 15
    if has_primary_obgyn_tax:
        score += 15
    if has_general_obgyn_code:
        score += 10
    if clean_county_geocode:
        score += 10
    if not_po_box:
        score += 5

    # Caps and penalties
    if mailing_only and not has_ga_primary_practice and not has_ga_secondary_practice:
        score = min(score, 49)   # cap at Medium
    if oos_practice_with_ga_license and not has_ga_primary_practice:
        score = min(score, 69)
    if failed_geocode:
        score = max(0, score - 10)
    if is_po_box_addr:
        score = max(0, score - 10)
    if multiple_conflicting_states:
        score = max(0, score - 15)

    return min(100, max(0, score))


def confidence_label(score: int) -> str:
    if score >= 85:
        return "Very High"
    if score >= 70:
        return "High"
    if score >= 50:
        return "Medium"
    if score >= 25:
        return "Low"
    return "Exclude / Very Low"


# ---------------------------------------------------------------------------
# 5.  Read NPPES CSV in chunks and filter to GA OB/GYN candidates
# ---------------------------------------------------------------------------

def process_nppes_zip(zip_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Stream main NPPES CSV in chunks, filter to GA OB/GYN candidates.
    Returns (providers_df, locations_df).
    """
    print("\n[STEP] Listing ZIP contents …")
    names = list_zip_contents(zip_path)
    main_csv = find_main_nppes_csv(names)
    pl_csv   = find_pl_csv(names)

    if not main_csv:
        raise FileNotFoundError("Could not identify main NPPES CSV inside ZIP.")

    print(f"  Main provider file : {main_csv}")
    print(f"  Practice loc file  : {pl_csv or '(not found — will skip)'}")

    # ---- Estimate row count for progress bar ----
    print("\n[STEP] Opening ZIP and reading main CSV in chunks …")
    candidate_chunks = []
    total_rows       = 0
    matched_rows     = 0

    with zipfile.ZipFile(zip_path, "r") as zf:
        with zf.open(main_csv) as raw_file:
            # Read header first to build column sets
            header_chunk = pd.read_csv(
                raw_file, nrows=0, dtype=str, low_memory=False,
                encoding="latin-1"
            )
            all_cols  = list(header_chunk.columns)
            tax_pairs = build_taxonomy_cols(all_cols)
            lic_pairs = build_license_cols(all_cols)

        # Re-open for chunked reading (ZipFile entry can only be read once)
        with zf.open(main_csv) as raw_file:
            reader = pd.read_csv(
                raw_file,
                chunksize=CHUNK_SIZE,
                dtype=str,
                low_memory=False,
                encoding="latin-1",
            )
            with tqdm(desc="  Scanning NPPES rows", unit=" rows", mininterval=2) as pbar:
                for chunk in reader:
                    chunk_size = len(chunk)
                    total_rows += chunk_size
                    pbar.update(chunk_size)

                    # --- Filter 1: Individual providers only (Entity Type = 1) ---
                    chunk = chunk[chunk.get("NPI Type", chunk.get("Entity Type Code", pd.Series(dtype=str))).fillna("") == "1"]
                    if chunk.empty:
                        continue

                    # --- Filter 2: Any OB/GYN taxonomy ---
                    tax_code_cols = [c for c, _ in tax_pairs]
                    has_obgyn_mask = pd.Series(False, index=chunk.index)
                    for col in tax_code_cols:
                        if col in chunk.columns:
                            has_obgyn_mask |= chunk[col].fillna("").str.startswith("207V")
                    chunk = chunk[has_obgyn_mask]
                    if chunk.empty:
                        continue

                    # --- Filter 3: Any GA evidence ---
                    ga_mask = pd.Series(False, index=chunk.index)

                    prac_state_col = None
                    mail_state_col = None
                    for candidate in [
                        "Provider Business Practice Location Address State Name",
                        "Provider Business Practice Location Address State Code",
                    ]:
                        if candidate in chunk.columns:
                            prac_state_col = candidate
                            break
                    for candidate in [
                        "Provider Business Mailing Address State Name",
                        "Provider Business Mailing Address State Code",
                    ]:
                        if candidate in chunk.columns:
                            mail_state_col = candidate
                            break

                    if prac_state_col:
                        ga_mask |= chunk[prac_state_col].fillna("").str.upper().isin(["GA", "GEORGIA"])
                    if mail_state_col:
                        ga_mask |= chunk[mail_state_col].fillna("").str.upper().isin(["GA", "GEORGIA"])

                    lic_state_cols = [sc for _, sc in lic_pairs if sc in chunk.columns]
                    for sc in lic_state_cols:
                        ga_mask |= chunk[sc].fillna("").str.upper() == "GA"

                    chunk = chunk[ga_mask]
                    if chunk.empty:
                        continue

                    matched_rows += len(chunk)
                    candidate_chunks.append(chunk)

    print(f"\n  Total rows scanned   : {total_rows:,}")
    print(f"  GA OB/GYN candidates : {matched_rows:,}")

    if not candidate_chunks:
        raise ValueError("No GA OB/GYN candidates found. Check taxonomy codes and state filters.")

    providers_df = pd.concat(candidate_chunks, ignore_index=True)

    # ---- Practice Location Reference file ----
    locations_df = pd.DataFrame()
    if pl_csv:
        print(f"\n[STEP] Reading Practice Location Reference file: {pl_csv}")
        with zipfile.ZipFile(zip_path, "r") as zf:
            try:
                with zf.open(pl_csv) as f:
                    locations_df = pd.read_csv(f, dtype=str, low_memory=False, encoding="latin-1")
                # Filter to GA only
                for sc in locations_df.columns:
                    if "state" in sc.lower():
                        locations_df = locations_df[
                            locations_df[sc].fillna("").str.upper().isin(["GA", "GEORGIA"])
                        ]
                        break
                print(f"  Practice location rows (GA): {len(locations_df):,}")
            except Exception as e:
                print(f"  [WARN] Could not read practice location file: {e}")

    return providers_df, locations_df


# ---------------------------------------------------------------------------
# 6.  Build enriched provider records
# ---------------------------------------------------------------------------

def enrich_providers(providers_df: pd.DataFrame) -> pd.DataFrame:
    """
    For each candidate provider row, compute:
    - taxonomies list
    - address fields
    - GA evidence flags
    - confidence score
    Returns an enriched DataFrame with one row per NPI.
    """
    cols = list(providers_df.columns)
    tax_pairs = build_taxonomy_cols(cols)
    lic_pairs = build_license_cols(cols)

    # Map column names defensively
    def get_col(df, candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    npi_col        = get_col(providers_df, ["NPI"])
    entity_col     = get_col(providers_df, ["NPI Type", "Entity Type Code"])
    fn_col         = get_col(providers_df, ["Provider First Name", "Provider First Name (Legal Name)"])
    ln_col         = get_col(providers_df, ["Provider Last Name (Legal Name)", "Provider Last Name"])
    cred_col       = get_col(providers_df, ["Provider Credential Text"])
    phone_col      = get_col(providers_df, [
        "Provider Business Practice Location Address Telephone Number",
        "Provider Business Practice Location Address Fax Number",
    ])

    prac_addr1_col = get_col(providers_df, ["Provider First Line Business Practice Location Address"])
    prac_addr2_col = get_col(providers_df, ["Provider Second Line Business Practice Location Address"])
    prac_city_col  = get_col(providers_df, ["Provider Business Practice Location Address City Name"])
    prac_state_col = get_col(providers_df, [
        "Provider Business Practice Location Address State Name",
        "Provider Business Practice Location Address State Code",
    ])
    prac_zip_col   = get_col(providers_df, ["Provider Business Practice Location Address Postal Code"])

    mail_addr1_col = get_col(providers_df, ["Provider First Line Business Mailing Address"])
    mail_city_col  = get_col(providers_df, ["Provider Business Mailing Address City Name"])
    mail_state_col = get_col(providers_df, [
        "Provider Business Mailing Address State Name",
        "Provider Business Mailing Address State Code",
    ])
    mail_zip_col   = get_col(providers_df, ["Provider Business Mailing Address Postal Code"])

    deact_col = get_col(providers_df, ["NPI Deactivation Date", "NPI Deactivation Reason Code"])

    records = []
    seen_npis = {}   # NPI -> index of best record in `records`

    for _, row in tqdm(providers_df.iterrows(), total=len(providers_df),
                       desc="  Enriching providers", mininterval=2):

        npi = str(row.get(npi_col, "") or "").strip()
        if not npi or npi == "nan":
            continue

        taxonomies   = extract_taxonomies(row, tax_pairs)
        tax_codes    = [c for c, _ in taxonomies]
        has_primary  = any(ip for _, ip in taxonomies if _ in OBGYN_TAXONOMY_CODES for _, ip in [(_, ip)])
        # redo cleanly
        primary_tax   = next((c for c, ip in taxonomies if ip), None)
        has_primary_obgyn = primary_tax in OBGYN_TAXONOMY_CODES if primary_tax else False
        has_general_obgyn = GENERAL_OBGYN_CODE in tax_codes

        # Practice address
        prac_addr   = str(row.get(prac_addr1_col, "") or "").strip()
        prac_addr2  = str(row.get(prac_addr2_col, "") or "").strip()
        full_prac   = (prac_addr + " " + prac_addr2).strip() if prac_addr2 else prac_addr
        prac_city   = str(row.get(prac_city_col, "") or "").strip()
        prac_state  = str(row.get(prac_state_col, "") or "").strip().upper()
        prac_zip    = str(row.get(prac_zip_col, "") or "").strip()[:5]
        if prac_zip.lower() == "nan":
            prac_zip = ""

        # Mailing address
        mail_state  = str(row.get(mail_state_col, "") or "").strip().upper()
        mail_city   = str(row.get(mail_city_col, "") or "").strip()
        mail_addr   = str(row.get(mail_addr1_col, "") or "").strip()
        mail_zip    = str(row.get(mail_zip_col, "") or "").strip()[:5]
        if mail_zip.lower() == "nan":
            mail_zip = ""

        # GA flags
        ga_prac   = prac_state in ("GA", "GEORGIA")
        ga_mail   = mail_state in ("GA", "GEORGIA")
        ga_lic    = has_ga_license(row, lic_pairs)
        ga_lic_obgyn = ga_license_on_obgyn(row, lic_pairs, taxonomies)
        mailing_only = ga_mail and not ga_prac

        oos_with_lic = (not ga_prac) and ga_lic

        # PO Box / address quality
        po_box    = is_po_box(full_prac) or is_po_box(mail_addr)
        not_po_box_flag = not po_box and bool(full_prac)

        fn = str(row.get(fn_col, "") or "").strip()
        ln = str(row.get(ln_col, "") or "").strip()
        cred = str(row.get(cred_col, "") or "").strip()
        phone = str(row.get(phone_col, "") or "").strip()

        # Deactivation check (informational)
        deact_raw = str(row.get(deact_col, "") or "").strip() if deact_col else ""
        is_deactivated = bool(deact_raw and deact_raw.lower() not in ("nan", "", "none"))

        # Build record — geocode will be filled in later
        rec = {
            "NPI": npi,
            "First Name": fn,
            "Last Name": ln,
            "Full Name": f"{fn} {ln}".strip(),
            "Credential": cred,
            "Phone": phone,
            "Practice Address Line 1": prac_addr,
            "Practice Address Line 2": prac_addr2,
            "Practice City": prac_city,
            "Practice State": prac_state,
            "Practice ZIP": prac_zip,
            "Mailing State": mail_state,
            "Mailing City": mail_city,
            "Mailing Address": mail_addr,
            "Mailing ZIP": mail_zip,
            "All Taxonomies": "|".join(tax_codes),
            "Primary Taxonomy": primary_tax or "",
            "General OBGYN Code Present": has_general_obgyn,
            "Has Primary OBGYN Taxonomy": has_primary_obgyn,
            "GA Practice Address": ga_prac,
            "GA Mailing Address": ga_mail,
            "GA License": ga_lic,
            "GA License on OBGYN Taxonomy": ga_lic_obgyn,
            "Mailing Only": mailing_only,
            "OOS Practice with GA License": oos_with_lic,
            "PO Box Address": po_box,
            "Is Deactivated (informational)": is_deactivated,
            # Workforce exclusion flags (HRSA methodology)
            "Likely Resident/Trainee": is_likely_resident(prac_addr, cred, prac_addr2),
            "Non-Physician Provider": is_non_physician(cred, primary_tax or ""),
            # Geocode fields — filled later
            "County": "",
            "Rural Status": "",
            "Geocode Match": "",
            "Geocode Matched Address": "",
            "Geocode Source": "",
            "Confidence Score": 0,
            "Confidence Label": "",
            "Multiple GA Counties": False,
            "All GA Counties Found": "",
            "Review Flag": "",
        }

        # Handle multiple NPIs: keep best (GA practice > GA mailing > other)
        if npi in seen_npis:
            prev_idx = seen_npis[npi]
            prev = records[prev_idx]
            # Prefer GA practice address over mailing only
            if ga_prac and not prev["GA Practice Address"]:
                records[prev_idx] = rec
            # Else keep existing
        else:
            seen_npis[npi] = len(records)
            records.append(rec)

    enriched = pd.DataFrame(records)
    print(f"  Unique NPI records: {len(enriched):,}")

    # Flag duplicate full names (same name, different NPI — possible same person)
    name_counts = enriched["Full Name"].value_counts()
    dup_names   = set(name_counts[name_counts > 1].index)
    enriched["Duplicate Name Flag"] = enriched["Full Name"].isin(dup_names)

    # Flag known same-address duplicates for manual review
    addr_counts = enriched["Practice Address Line 1"].str.upper().value_counts()
    dup_addrs   = set(addr_counts[addr_counts > 1].index)
    enriched["Duplicate Address Flag"] = enriched["Practice Address Line 1"].str.upper().isin(dup_addrs)

    n_dup_names = enriched["Duplicate Name Flag"].sum()
    n_dup_addr  = enriched["Duplicate Address Flag"].sum()
    print(f"  Duplicate name flags  : {n_dup_names}")
    print(f"  Duplicate address flags: {n_dup_addr}")
    return enriched


# ---------------------------------------------------------------------------
# 7.  Geocoding pass
# ---------------------------------------------------------------------------

def geocode_providers(enriched: pd.DataFrame) -> pd.DataFrame:
    """
    Geocode distinct practice addresses. Uses cache; calls Census API for misses.
    """
    cache = load_geocode_cache()
    new_calls = 0

    # Build list of unique non-empty GA practice addresses to geocode
    ga_prac_mask = enriched["GA Practice Address"] == True  # noqa: E712
    rows_to_geo  = enriched[ga_prac_mask].copy()

    print(f"\n[STEP] Geocoding {len(rows_to_geo):,} GA practice address rows …")

    counties   = []
    match_info = []
    match_addr = []
    sources    = []

    for _, row in tqdm(rows_to_geo.iterrows(), total=len(rows_to_geo),
                       desc="  Geocoding", mininterval=3):
        street  = row["Practice Address Line 1"]
        city    = row["Practice City"]
        state   = row["Practice State"]
        zipcode = row["Practice ZIP"]

        key = normalize_address(street, city, state, zipcode)

        if key in cache:
            result = cache[key]
            src = "cache"
        elif is_po_box(street) or not street:
            result = {"matched": False, "county": None, "match_type": "PO Box / no street",
                      "matched_address": None}
            src = "skip"
        else:
            result = geocode_census(street, city, state, zipcode)
            cache[key] = result
            new_calls += 1
            src = "census"
            # Be polite
            time.sleep(0.15)

        county_raw = result.get("county") or ""
        county_clean = re.sub(r"\s+County$", "", county_raw, flags=re.IGNORECASE).strip()

        counties.append(county_clean)
        match_info.append(result.get("match_type", ""))
        match_addr.append(result.get("matched_address", "") or "")
        sources.append(src)

    rows_to_geo["County"] = counties
    rows_to_geo["Geocode Match"] = match_info
    rows_to_geo["Geocode Matched Address"] = match_addr
    rows_to_geo["Geocode Source"] = sources

    # Merge back
    enriched.update(rows_to_geo[["County", "Geocode Match", "Geocode Matched Address", "Geocode Source"]])

    # ZIP fallback for failed geocodes with GA practice address
    for idx, row in enriched[ga_prac_mask].iterrows():
        if not row["County"] and row["Practice ZIP"]:
            z = str(row["Practice ZIP"]).strip()[:5]
            if z in GA_ZIP_TO_COUNTY:
                enriched.at[idx, "County"] = GA_ZIP_TO_COUNTY[z]
                enriched.at[idx, "Geocode Source"] = "zip_fallback"

    # Rural classification
    enriched["Rural Status"] = enriched["County"].apply(classify_rural)

    if new_calls > 0:
        save_geocode_cache(cache)
        print(f"  Geocode cache updated ({new_calls:,} new calls saved).")

    return enriched


# ---------------------------------------------------------------------------
# 8.  Confidence scoring pass
# ---------------------------------------------------------------------------

def score_providers(enriched: pd.DataFrame) -> pd.DataFrame:
    scores = []
    labels = []
    flags  = []

    for _, row in enriched.iterrows():
        clean_geo = bool(row["County"] and row["Geocode Source"] not in ("", "skip")
                         and "Error" not in str(row["Geocode Match"]))

        score = compute_confidence(
            has_ga_primary_practice    = bool(row["GA Practice Address"]),
            has_ga_secondary_practice  = False,   # PL file handled separately
            ga_license_on_obgyn_tax    = bool(row["GA License on OBGYN Taxonomy"]),
            has_any_ga_license         = bool(row["GA License"]),
            has_primary_obgyn_tax      = bool(row["Has Primary OBGYN Taxonomy"]),
            has_general_obgyn_code     = bool(row["General OBGYN Code Present"]),
            clean_county_geocode       = clean_geo,
            not_po_box                 = not bool(row["PO Box Address"]),
            mailing_only               = bool(row["Mailing Only"]),
            failed_geocode             = not clean_geo and bool(row["GA Practice Address"]),
            is_po_box_addr             = bool(row["PO Box Address"]),
            multiple_conflicting_states= False,
            oos_practice_with_ga_license = bool(row["OOS Practice with GA License"]),
        )
        label = confidence_label(score)

        # Build review flag
        flag_parts = []
        if not row["County"]:
            flag_parts.append("Failed geocode")
        if row["PO Box Address"]:
            flag_parts.append("PO Box")
        if row["Mailing Only"]:
            flag_parts.append("Mailing address only")
        if row["OOS Practice with GA License"]:
            flag_parts.append("OOS practice / GA license only")
        if row["Is Deactivated (informational)"]:
            flag_parts.append("NPI deactivation on record")
        if row["Rural Status"] == "Partial rural - tract/address review needed":
            flag_parts.append("Partial-rural county")
        if row["Rural Status"] == "Unknown - geocode/manual review":
            flag_parts.append("Unknown rural status")
        if label in ("Low", "Exclude / Very Low"):
            flag_parts.append(f"Low confidence ({score})")

        scores.append(score)
        labels.append(label)
        flags.append("; ".join(flag_parts) if flag_parts else "")

    enriched["Confidence Score"] = scores
    enriched["Confidence Label"] = labels
    enriched["Review Flag"]      = flags
    return enriched


# ---------------------------------------------------------------------------
# 9.  Cohort assignment
# ---------------------------------------------------------------------------

def assign_cohorts(enriched: pd.DataFrame) -> pd.DataFrame:
    """
    Add cohort membership columns aligned with HRSA HWSM methodology:
    - In_Umbrella: all 207V OB/GYN taxonomy providers with GA practice evidence
    - In_Core_General: attending physician OB/GYNs only — excludes residents,
      non-physician providers, and wrong-taxonomy records per HRSA definition
      of 'active in the workforce providing direct patient care'
    - In_APP_Cohort: non-physician OB/GYN workforce (CNMs, NPs, PAs) — separate
      count, not included in physician estimates
    - In_Resident_Cohort: likely residents/trainees — tracked separately as
      pipeline indicator, not counted as active workforce
    """

    # Exclusion flags
    is_resident    = enriched["Likely Resident/Trainee"].astype(bool)
    is_non_phys    = enriched["Non-Physician Provider"].astype(bool)
    is_wrong_tax   = enriched["Primary Taxonomy"].isin(NON_OBGYN_PRIMARY_TAXONOMIES - {"207VX0000X"})

    # Umbrella: any 207V taxonomy + GA practice evidence + not wrong taxonomy
    umbrella = (
        (enriched["GA Practice Address"] | (enriched["GA License"] & ~enriched["GA Practice Address"]))
        & enriched["Confidence Score"].ge(25)
        & ~is_wrong_tax
    )
    enriched["In_Umbrella"] = umbrella

    # Check if all taxonomy codes are subspecialist-only
    def all_subspecialist(tax_str: str) -> bool:
        codes = [c.strip() for c in str(tax_str).split("|") if c.strip()]
        if not codes:
            return True
        return all(c in SUBSPECIALIST_ONLY_CODES for c in codes)

    enriched["All Subspecialist Only"] = enriched["All Taxonomies"].apply(all_subspecialist)

    is_general = enriched["General OBGYN Code Present"] | (
        enriched["Primary Taxonomy"] == GENERAL_OBGYN_CODE
    )

    # Core General Physician OB/GYN cohort — HRSA-aligned attending physician count
    # Excludes: residents/trainees, non-physician providers, wrong taxonomy
    core_general = (
        umbrella
        & (is_general | ~enriched["All Subspecialist Only"])
        & enriched["GA Practice Address"]
        & enriched["Confidence Score"].ge(25)
        & ~is_resident      # HRSA: excludes residents
        & ~is_non_phys      # HRSA: physician-only count
        & ~is_wrong_tax     # taxonomy must be OB/GYN family
    )
    enriched["In_Core_General"] = core_general

    # Advanced Practice Provider cohort (CNMs, NPs, PAs) — separate workforce count
    app_cohort = (
        enriched["GA Practice Address"]
        & is_non_phys
        & ~is_wrong_tax
        & enriched["Confidence Score"].ge(25)
    )
    enriched["In_APP_Cohort"] = app_cohort

    # Resident/trainee cohort — pipeline indicator
    enriched["In_Resident_Cohort"] = (
        is_resident
        & enriched["GA Practice Address"]
        & ~is_non_phys
    )

    # Log exclusion counts
    n_res  = is_resident.sum()
    n_app  = is_non_phys.sum()
    n_tax  = is_wrong_tax.sum()
    print(f"  Excluded as likely residents/trainees : {n_res}")
    print(f"  Separated as non-physician APPs       : {n_app}")
    print(f"  Excluded for wrong primary taxonomy   : {n_tax}")

    return enriched


# ---------------------------------------------------------------------------
# 10.  DuckDB store
# ---------------------------------------------------------------------------

def save_to_duckdb(enriched: pd.DataFrame):
    """Persist enriched provider table to DuckDB for ad-hoc queries."""
    print(f"\n[STEP] Saving to DuckDB: {DUCKDB_PATH}")
    con = duckdb.connect(str(DUCKDB_PATH))
    con.execute("DROP TABLE IF EXISTS ga_obgyn_providers")
    con.execute("CREATE TABLE ga_obgyn_providers AS SELECT * FROM enriched")
    con.close()
    print("  DuckDB saved.")


# ---------------------------------------------------------------------------
# 11.  Excel output helpers
# ---------------------------------------------------------------------------

def auto_width(ws, max_width=60):
    """Set reasonable column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def df_to_sheet(ws, df: pd.DataFrame, table_name: str = None):
    """Write a DataFrame to a worksheet as an Excel table."""
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    freeze_and_filter(ws)
    auto_width(ws)
    if table_name and len(df) > 0:
        from openpyxl.worksheet.table import Table, TableStyleInfo
        ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
        tbl = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)


def write_clean_workbook(enriched: pd.DataFrame, out_path: Path):
    print(f"\n[STEP] Writing clean workbook: {out_path}")
    wb = Workbook()

    # ---- Core General cohort subset ----
    core = enriched[enriched["In_Core_General"]].copy()
    # Remove duplicate NPIs (should not exist, but safety check)
    core = core.drop_duplicates(subset=["NPI"])

    umbrella = enriched[enriched["In_Umbrella"]].copy().drop_duplicates(subset=["NPI"])

    # ---- Tab 1: Start Here ----
    ws_start = wb.active
    ws_start.title = "Start Here"

    hdr_fill = PatternFill("solid", fgColor="003366")
    hdr_font = Font(color="FFFFFF", bold=True, size=12)

    def add_header(ws, text, row):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=13, color="003366")

    def add_row(ws, label, value, row):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)

    rural_counts = core["Rural Status"].value_counts().to_dict()
    total_core = len(core)

    r = 1
    ws_start.cell(r, 1, "Georgia OB/GYN Practice Estimate (NPI/NPPES-Based)").font = Font(bold=True, size=16)
    r += 1
    ws_start.cell(r, 1, f"Generated: {date.today().strftime('%B %d, %Y')}").font = Font(italic=True, color="666666")
    r += 2

    add_header(ws_start, "RECOMMENDED ESTIMATE — Core General OB/GYN Cohort", r); r += 1
    add_row(ws_start, "Total Core General OB/GYN Physicians", total_core, r); r += 1
    add_row(ws_start, "  — Rural counties", rural_counts.get("Rural", 0), r); r += 1
    add_row(ws_start, "  — Partial rural (tract/address review needed)", rural_counts.get("Partial rural - tract/address review needed", 0), r); r += 1
    add_row(ws_start, "  — Not rural", rural_counts.get("Not rural", 0), r); r += 1
    add_row(ws_start, "  — Unknown / manual review", rural_counts.get("Unknown - geocode/manual review", 0), r); r += 1
    r += 1

    add_header(ws_start, "ADDITIONAL COUNTS", r); r += 1
    add_row(ws_start, "Core Umbrella OB/GYN (all 207V subspecialties)", len(umbrella), r); r += 1
    r += 1

    add_header(ws_start, "METHODOLOGY CAVEAT", r); r += 1
    caveat = (
        "This estimate is derived from the CMS/NPPES National Provider Identifier (NPI) registry. "
        "It is an evidence-weighted estimate, not a census. Providers may have stale, incomplete, "
        "or administrative NPI records. The recommended point estimate includes only individual "
        "physicians with a Georgia practice address, general OB/GYN taxonomy evidence, and "
        "at least Low confidence. Rural classification is county-level and may require "
        "tract/address-level review for partial-rural counties."
    )
    ws_start.cell(r, 1, caveat)
    ws_start.cell(r, 1).alignment = Alignment(wrap_text=True)
    ws_start.row_dimensions[r].height = 72
    r += 2

    add_row(ws_start, "Data Source", "CMS/NPPES Full Replacement Monthly File", r); r += 1
    add_row(ws_start, "Primary Filter", "Individual NPI, 207V taxonomy family, Georgia practice/license evidence", r); r += 1

    ws_start.column_dimensions["A"].width = 52
    ws_start.column_dimensions["B"].width = 32

    # ---- Tab 2: Core General OB/GYNs ----
    ws_core = wb.create_sheet("Core General OBGYNs")
    display_cols = [
        "NPI", "Full Name", "Credential",
        "Practice Address Line 1", "Practice City", "Practice State", "Practice ZIP",
        "County", "Rural Status", "Phone",
        "Primary Taxonomy", "General OBGYN Code Present",
        "GA License", "Confidence Score", "Confidence Label", "Review Flag",
    ]
    existing = [c for c in display_cols if c in core.columns]
    df_to_sheet(ws_core, core[existing], "CoreGeneralOBGYN")

    # ---- Tab 3: County Counts ----
    ws_county = wb.create_sheet("County Counts")
    county_grp = (
        core.groupby(["County", "Rural Status"])
        .size()
        .reset_index(name="Core General OB/GYN Count")
    )
    county_grp["Share of State Total"] = (county_grp["Core General OB/GYN Count"] / total_core).map(
        lambda x: f"{x:.1%}" if total_core > 0 else "N/A"
    )
    county_grp = county_grp.sort_values("Core General OB/GYN Count", ascending=False)
    df_to_sheet(ws_county, county_grp, "CountyCounts")

    # ---- Tab 4: Rural Counts ----
    ws_rural = wb.create_sheet("Rural Counts")
    rural_order = ["Rural", "Partial rural - tract/address review needed",
                   "Not rural", "Unknown - geocode/manual review"]
    rural_df = (
        core["Rural Status"]
        .value_counts()
        .reindex(rural_order, fill_value=0)
        .reset_index()
    )
    rural_df.columns = ["Rural Status", "Count"]
    rural_df["Pct of Core General"] = (rural_df["Count"] / total_core).map(
        lambda x: f"{x:.1%}" if total_core > 0 else "N/A"
    )
    df_to_sheet(ws_rural, rural_df, "RuralCounts")

    # ---- Tab 5: Manual Review ----
    ws_review = wb.create_sheet("Manual Review")
    review_mask = (
        enriched["Review Flag"].str.len().gt(0)
        | enriched["Confidence Label"].isin(["Low", "Exclude / Very Low", "Medium"])
        | enriched["Mailing Only"]
        | enriched["PO Box Address"]
    )
    review_df = enriched[review_mask].drop_duplicates(subset=["NPI"])
    df_to_sheet(ws_review, review_df[[c for c in display_cols if c in review_df.columns] + ["Review Flag"]], "ManualReview")

    # ---- Tab 6: Methodology Notes ----
    ws_method = wb.create_sheet("Methodology Notes")
    notes = [
        ("Data Source", "CMS/NPPES Full Replacement Monthly File (https://download.cms.gov/nppes/NPI_Files.html)"),
        ("Taxonomy Scope", "Individual providers (Entity Type = 1) with any 207V-family OB/GYN taxonomy code"),
        ("Included Codes", ", ".join(f"{k} ({v})" for k, v in OBGYN_TAXONOMY_CODES.items())),
        ("Georgia Evidence", "Primary practice address state = GA, or any license state code = GA"),
        ("Address Hierarchy", "1. Primary practice address (GA) > 2. Secondary GA practice location > 3. GA mailing address (weak signal only)"),
        ("Geocoding", "U.S. Census Geocoder (single-address API). Results cached locally. ZIP fallback if geocode fails and ZIP is unambiguous."),
        ("Rural Classification", "County-level. Partial-rural counties flagged for tract/address review. Taxonomy: Rural / Partial rural / Not rural / Unknown."),
        ("Confidence Scoring", "Evidence-weighted 0–100 score. GA practice address (+35), secondary GA location (+25), GA OB/GYN license (+25), GA any license (+15), primary OB/GYN taxonomy (+15), general OB/GYN code (+10), clean geocode (+10), non-PO Box (+5). Penalties/caps for mailing-only, OOS practice, PO Box, failed geocode."),
        ("Core General Cohort", "GA practice address, general OB/GYN taxonomy evidence (207V00000X or primary = OB/GYN), not exclusively subspecialist-only, confidence >= Low"),
        ("Core Umbrella Cohort", "GA practice address or GA license, any 207V taxonomy, confidence >= Low"),
        ("Deduplication", "One row per NPI in provider-level counts. All locations preserved in audit workbook."),
        ("LIMITATION", (
            "This workbook is an NPI/NPPES-derived, evidence-weighted estimate of Georgia OB/GYN "
            "practice locations. NPI/NPPES data is not a perfect census. Provider records may be "
            "stale, incomplete, miscoded, or associated with administrative, mailing, or outdated "
            "practice locations. Some physicians may have moved, retired, changed specialties, "
            "stopped practicing clinically, or failed to update their NPI record. Conversely, some "
            "active physicians may be missed if their taxonomy, address, license, or NPI information "
            "is incomplete or inconsistent. The recommended estimate should be interpreted as a "
            "high-confidence public-data estimate, not a definitive licensure, claims-based, "
            "employment-based, or clinically active workforce count."
        )),
    ]
    for i, (label, text) in enumerate(notes, start=1):
        ws_method.cell(i, 1, label).font = Font(bold=True)
        ws_method.cell(i, 2, text).alignment = Alignment(wrap_text=True)
        ws_method.row_dimensions[i].height = max(30, min(120, len(str(text)) // 3))
    ws_method.column_dimensions["A"].width = 28
    ws_method.column_dimensions["B"].width = 100

    wb.save(out_path)
    print(f"  Saved: {out_path}")


def write_audit_workbook(enriched: pd.DataFrame, locations_df: pd.DataFrame, out_path: Path):
    print(f"\n[STEP] Writing audit workbook: {out_path}")
    wb = Workbook()

    # Tab 1: All Provider Records
    ws1 = wb.active
    ws1.title = "All Provider Records"
    df_to_sheet(ws1, enriched, "AllProviders")

    # Tab 2: All Practice Locations
    ws2 = wb.create_sheet("All Practice Locations")
    loc_cols = [
        "NPI", "Full Name", "Practice Address Line 1", "Practice City",
        "Practice State", "Practice ZIP", "County", "Rural Status",
        "Geocode Source", "Geocode Match", "Geocode Matched Address",
    ]
    existing = [c for c in loc_cols if c in enriched.columns]
    df_to_sheet(ws2, enriched[existing], "AllLocations")

    # Tab 3: All Taxonomies
    ws3 = wb.create_sheet("All Taxonomies")
    tax_rows = []
    for _, row in enriched.iterrows():
        for code in str(row.get("All Taxonomies", "")).split("|"):
            code = code.strip()
            if code:
                tax_rows.append({
                    "NPI": row["NPI"],
                    "Full Name": row.get("Full Name", ""),
                    "Taxonomy Code": code,
                    "Taxonomy Description": OBGYN_TAXONOMY_CODES.get(code, code),
                    "Is Primary": code == row.get("Primary Taxonomy", ""),
                    "General OBGYN Code": code == GENERAL_OBGYN_CODE,
                })
    df_to_sheet(ws3, pd.DataFrame(tax_rows), "AllTaxonomies")

    # Tab 4: Geocode Diagnostics
    ws4 = wb.create_sheet("Geocode Diagnostics")
    failed_geo = enriched[
        enriched["GA Practice Address"] & (enriched["County"].fillna("") == "")
    ].copy()
    df_to_sheet(ws4, failed_geo[[c for c in [
        "NPI", "Full Name", "Practice Address Line 1", "Practice City",
        "Practice State", "Practice ZIP", "Geocode Match", "Geocode Source",
        "PO Box Address", "Review Flag",
    ] if c in failed_geo.columns]], "GeocodeDiag")

    # Tab 5: License Verification Queue
    ws5 = wb.create_sheet("License Verification Queue")
    lic_q = enriched[enriched["GA License"] & ~enriched["GA Practice Address"]].copy()
    df_to_sheet(ws5, lic_q[[c for c in [
        "NPI", "Full Name", "Credential", "Practice State",
        "Mailing State", "GA License", "Confidence Score", "Confidence Label",
    ] if c in lic_q.columns]], "LicenseQueue")

    # Tab 6: Excluded Records
    ws6 = wb.create_sheet("Excluded Records")
    excluded = enriched[enriched["Confidence Label"].isin(["Exclude / Very Low"])].copy()
    df_to_sheet(ws6, excluded, "ExcludedRecords")

    # Tab 7: Cohort Logic Trace
    ws7 = wb.create_sheet("Cohort Logic Trace")
    trace_cols = [
        "NPI", "Full Name", "GA Practice Address", "GA License",
        "GA License on OBGYN Taxonomy", "General OBGYN Code Present",
        "Has Primary OBGYN Taxonomy", "All Subspecialist Only",
        "Mailing Only", "PO Box Address", "Is Deactivated (informational)",
        "Confidence Score", "Confidence Label",
        "In_Umbrella", "In_Core_General", "Review Flag",
    ]
    existing = [c for c in trace_cols if c in enriched.columns]
    df_to_sheet(ws7, enriched[existing], "CohortTrace")

    wb.save(out_path)
    print(f"  Saved: {out_path}")


# ---------------------------------------------------------------------------
# 12.  Validation checks
# ---------------------------------------------------------------------------

def run_validation(enriched: pd.DataFrame, total_rows_scanned: int, matched_rows: int,
                   skipped_raw: bool):
    print("\n" + "="*60)
    print("VALIDATION SUMMARY")
    print("="*60)

    core    = enriched[enriched["In_Core_General"]]
    umbrella = enriched[enriched["In_Umbrella"]]
    rural_vc = core["Rural Status"].value_counts()

    total_npi       = len(enriched)
    total_obgyn_tax = len(enriched[enriched["All Taxonomies"].str.len().gt(0)])
    total_ga_prac   = len(enriched[enriched["GA Practice Address"]])
    n_core          = len(core.drop_duplicates("NPI"))
    n_umb           = len(umbrella.drop_duplicates("NPI"))
    n_app           = len(enriched[enriched.get("In_APP_Cohort", False)].drop_duplicates("NPI")) if "In_APP_Cohort" in enriched.columns else 0
    n_resident      = len(enriched[enriched.get("In_Resident_Cohort", False)].drop_duplicates("NPI")) if "In_Resident_Cohort" in enriched.columns else 0
    n_rural         = rural_vc.get("Rural", 0)
    n_partial       = rural_vc.get("Partial rural - tract/address review needed", 0)
    n_not_rural     = rural_vc.get("Not rural", 0)
    n_unknown       = rural_vc.get("Unknown - geocode/manual review", 0)
    n_failed_geo    = len(enriched[enriched["GA Practice Address"] & enriched["County"].eq("")])
    n_multi_county  = len(enriched[enriched["Multiple GA Counties"]])
    n_excluded      = len(enriched[enriched["Confidence Label"].isin(["Exclude / Very Low"])])

    print(f"  Total individual NPIs processed          : {total_npi:>8,}")
    print(f"  Total with any OB/GYN taxonomy           : {total_obgyn_tax:>8,}")
    print(f"  Total with GA practice-location evidence : {total_ga_prac:>8,}")
    print(f"  Core Umbrella OB/GYN count               : {n_umb:>8,}")
    print(f"  Core General OB/GYN (attending MDs/DOs) : {n_core:>8,}")
    print(f"    — Rural                                : {n_rural:>8,}")
    print(f"    — Partial rural                        : {n_partial:>8,}")
    print(f"    — Not rural                            : {n_not_rural:>8,}")
    print(f"    — Unknown / manual review              : {n_unknown:>8,}")
    print(f"  Advanced Practice Providers (APP cohort) : {n_app:>8,}  (CNMs, NPs, PAs — separate count)")
    print(f"  Likely residents/trainees (pipeline)     : {n_resident:>8,}  (excluded from workforce count)")
    print(f"  Failed geocodes                          : {n_failed_geo:>8,}")
    print(f"  Providers with multiple GA counties      : {n_multi_county:>8,}")
    print(f"  Records excluded (Exclude/Very Low)      : {n_excluded:>8,}")
    print(f"  Candidate rows loaded from NPPES         : {matched_rows:>8,}")
    print(f"  Total NPPES rows scanned                 : {total_rows_scanned:>8,}")
    print(f"  Raw NPPES processing skipped?            : {'YES' if skipped_raw else 'NO'}")

    # Internal consistency
    print("\n  Internal consistency checks:")
    ok = True
    if n_core > n_umb:
        print("  [FAIL] Core General > Core Umbrella — should not happen")
        ok = False
    else:
        print("  [OK]   Core General <= Core Umbrella")

    rural_sum = n_rural + n_partial + n_not_rural + n_unknown
    if rural_sum != n_core:
        print(f"  [WARN] Rural+Partial+NotRural+Unknown ({rural_sum}) != Core General ({n_core})")
        ok = False
    else:
        print("  [OK]   Rural/Partial/NotRural/Unknown sum equals Core General count")

    dup_npis = core["NPI"].duplicated().sum()
    if dup_npis > 0:
        print(f"  [FAIL] {dup_npis} duplicate NPIs in Core General sheet")
        ok = False
    else:
        print("  [OK]   No duplicate NPIs in Core General")

    if ok:
        print("\n  All checks passed.")
    print("="*60)


# ---------------------------------------------------------------------------
# 13.  Main
# ---------------------------------------------------------------------------

def ask_refresh() -> bool:
    """Ask user whether to refresh processed data or reuse."""
    try:
        ans = input(
            "\nProcessed files found. Reuse them? [Y/n]: "
        ).strip().lower()
        return ans == "n"
    except EOFError:
        # Non-interactive — default to reuse
        return False


def main():
    print("\n" + "="*60)
    print(" GA OB/GYN Rural NPI Lookup — v6 (NPPES Full File)")
    print("="*60)

    make_dirs()
    zip_path = find_nppes_zip()
    print(f"  NPPES ZIP: {zip_path.name}")

    # --- Decide whether to skip raw processing ---
    processed_ok = (
        PARQUET_PROVIDERS.exists()
        and PARQUET_PROVIDERS.stat().st_size > 0
    )
    skipped_raw = False

    if processed_ok:
        do_refresh = ask_refresh()
        if not do_refresh:
            print("\n[STEP] Loading from existing Parquet files …")
            enriched = pd.read_parquet(PARQUET_PROVIDERS)
            skipped_raw = True
            matched_rows = len(enriched)
            total_rows_scanned = -1   # unknown when skipping
        else:
            print("\n[STEP] Refreshing from raw NPPES ZIP …")
            providers_df, locations_df = process_nppes_zip(zip_path)
            matched_rows = len(providers_df)
            total_rows_scanned = matched_rows   # approximation; exact tally inside function

            print("\n[STEP] Enriching providers …")
            enriched = enrich_providers(providers_df)

            print(f"\n[STEP] Saving candidates to Parquet: {PARQUET_PROVIDERS}")
            enriched.to_parquet(PARQUET_PROVIDERS, index=False)

            if not locations_df.empty:
                locations_df.to_parquet(PARQUET_LOCATIONS, index=False)
    else:
        print("\n[STEP] No processed files found. Processing raw NPPES ZIP …")
        providers_df, locations_df = process_nppes_zip(zip_path)
        matched_rows = len(providers_df)
        total_rows_scanned = matched_rows

        print("\n[STEP] Enriching providers …")
        enriched = enrich_providers(providers_df)

        print(f"\n[STEP] Saving candidates to Parquet: {PARQUET_PROVIDERS}")
        enriched.to_parquet(PARQUET_PROVIDERS, index=False)
        if not locations_df.empty:
            locations_df.to_parquet(PARQUET_LOCATIONS, index=False)

    # Load locations for audit if available
    locations_df = pd.DataFrame()
    if PARQUET_LOCATIONS.exists():
        try:
            locations_df = pd.read_parquet(PARQUET_LOCATIONS)
        except Exception:
            pass

    # --- Geocode ---
    enriched = geocode_providers(enriched)

    # --- Score ---
    print("\n[STEP] Computing confidence scores …")
    enriched = score_providers(enriched)

    # --- Cohorts ---
    print("\n[STEP] Assigning cohorts …")
    enriched = assign_cohorts(enriched)

    # --- DuckDB ---
    save_to_duckdb(enriched)

    # --- Outputs ---
    core    = enriched[enriched["In_Core_General"]].drop_duplicates("NPI")
    umbrella = enriched[enriched["In_Umbrella"]].drop_duplicates("NPI")

    write_clean_workbook(enriched, OUT_DIR / "GA_OBGYN_Rural_v6_CLEAN.xlsx")
    write_audit_workbook(enriched, locations_df, OUT_DIR / "GA_OBGYN_Rural_v6_AUDIT.xlsx")

    core.to_csv(OUT_DIR / "GA_OBGYN_Rural_v6_Core_General.csv", index=False)
    umbrella.to_csv(OUT_DIR / "GA_OBGYN_Rural_v6_Core_Umbrella.csv", index=False)
    if "In_APP_Cohort" in enriched.columns:
        app = enriched[enriched["In_APP_Cohort"]].drop_duplicates("NPI")
        app.to_csv(OUT_DIR / "GA_OBGYN_Rural_v6_APP_Cohort.csv", index=False)
    if "In_Resident_Cohort" in enriched.columns:
        res = enriched[enriched["In_Resident_Cohort"]].drop_duplicates("NPI")
        res.to_csv(OUT_DIR / "GA_OBGYN_Rural_v6_Residents.csv", index=False)
    enriched.to_csv(OUT_DIR / "GA_OBGYN_Rural_v6_All_Provider_Audit.csv", index=False)

    # --- Validation ---
    run_validation(enriched, total_rows_scanned, matched_rows, skipped_raw)

    print("\nDone. Outputs written to:")
    for f in sorted(OUT_DIR.iterdir()):
        print(f"  {f}")


if __name__ == "__main__":
    main()
